# This script is used to get some stats
# for experiment 1 sample results

import sys
import os
import argparse
import math
import openpyxl
from itertools import chain, combinations, product
from collections import Counter
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.formatting.rule import ColorScaleRule

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.versioning import get_version
from utils.serialisation import load, save
from visualisation.row_data import RowData

def compute_stats(rowData, catsConst, catsVar, topPercentage, combsLength):
    filterDict = rowData.get_filter_dict()
    masterFilterDict = {k: v if v else filterDict[k] for k,v in catsVar.items()}
    if combsLength <= 0:
        combsLength = len(masterFilterDict) - 1
    # get combinations of filter dict
    combs = chain.from_iterable(combinations(masterFilterDict.keys(), r) for r in range(combsLength + 1))
    filterDicts = [{k: masterFilterDict[k] for k in c} for c in combs]

    result = []
    for filterDict in filterDicts:
        filterDict = {**catsConst, **filterDict}
        keys = filterDict.keys()
        values = filterDict.values()
        combs = list(product(*values))
        filters = [{key: [ value ] for key, value in zip(keys, comb)} for comb in combs]

        # Columns to get stats about
        colStatKeys = {s: RowData.HEADER.index(s) for s in masterFilterDict.keys() - filterDict.keys()}

        #filtered = []
        stats=[]
        for f in filters:
            fRows = rowData.filter_rows(f)
            totalRowCount = len(fRows)
            numItems = math.ceil(len(fRows) * topPercentage * 0.01)
            fRows = fRows[:numItems]
            stat = { s : Counter([row[i] for row in fRows]) for s, i in colStatKeys.items() }
            stat = { s : { k : [c, c / numItems * 100] for k, c in count.items() } for s, count in stat.items() }
            stats.append({'Filter': f, 'Stats': stat, 'RowCount': totalRowCount, 'RowsFiltered': numItems })
            #filtered.append(fRows)
        result.append({'Combination': filterDict, 'Permutations': stats})
    
    return {
        'ConstantCategories': catsConst,
        'VariableCategories': masterFilterDict,
        'TopPercentage': topPercentage,
        'CombinationLength': combsLength,
        'Result': result
    }

def to_excel(outputFile, results):
    workbook = openpyxl.Workbook()
    
    blue_fill = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
    # Define the color scale rule
    color_scale_rule = ColorScaleRule(
        start_type="num", start_value=0, start_color=Color(rgb="FF0000"),
        end_type="num", end_value=100, end_color=Color(rgb="00FF00")
    )

    # Sheet generator
    def sheet_enumerable():
        yield workbook.active
        while True:
            yield workbook.create_sheet()

    sheetGen = sheet_enumerable()
    for result in results:
        # Get sheet
        sheet = next(sheetGen)

        # Gather the const vars
        catsConst = result['ConstantCategories']
        percentage = result['TopPercentage']
        sheet.title = f"{'{:02}'.format(percentage)}_{'_'.join(v[0] for v in catsConst.values())}"

        # Form all header columns
        allComibnations = {**result['ConstantCategories'], **result['VariableCategories']}
        headerIndex = {}
        valueIndex = {}
        count = 0
        for category, values in allComibnations.items():
            headerIndex[category] = count
            for value in values:
                valueIndex[value] = count
                count += 1

        # Add header and make it bold
        row = [''] * len(valueIndex) + ['Stats']
        for k, v in headerIndex.items():
            row[v] = k
        sheet.append(row)
        sheet.append(list(valueIndex.keys()) + ['Filtered', 'Count'])
        for i in range(1, 3):
            for cell in sheet[i]:
                cell.font = Font(bold=True)

        resList = result['Result']
        for resItem in resList:
            # Optional, show header again

            perms = resItem['Permutations']
            for perm in perms:

                # Append results
                stats = perm['Stats']

                row = [''] * len(valueIndex) + [perm['RowsFiltered'], perm['RowCount']]

                sheet.append(row)
                s_row = sheet[sheet.max_row]

                # Fixed values are always 100%
                for key, stat in perm['Filter'].items():
                    cell = s_row[valueIndex[stat[0]]]
                    cell.value = 100
                    cell.fill = blue_fill

                # Fill frequency row
                for col, statDict in stats.items():
                    for key, stat in statDict.items():
                        cell = s_row[valueIndex[key]]
                        cell.value = round(stat[1], 2)
                        sheet.conditional_formatting.add(cell.coordinate, color_scale_rule)
        
            sheet.append([])
    
    workbook.save(outputFile)

def main():
    versionString = get_version().to_string()
    parser = argparse.ArgumentParser(description=f'Generate a matrix for top x% results.\n{versionString}')
    parser.add_argument('--result', default='all_results.json', help='The result to load')
    parser.add_argument('--percentages', default='5,10,15,25', help='The top percentages (comma separated) of filtered results that will be considered')
    parser.add_argument('--image-loaders', default='gray_tm', help='Which image loaders will be considered (comma separated) (tabbed)')
    parser.add_argument('--combinations', default=0, type=int, help='The largest selection of comibnations. 2 will give pair combinations from catsVar')
    parser.add_argument('--use-best-metrics', default=False, action='store_true', help='Sort by best metrics. There will be no filtering occuring using one metric')
    parser.add_argument('--output', default='stats', help='Output file name without extension')

    args = parser.parse_args()
    topPercentages = list(map(int, args.percentages.split(',')))
    useBestMetrics = args.use_best_metrics
    imageLoaders = list(map(lambda x: x.strip().lower(), args.image_loaders.split(',')))
    combsLength = args.combinations
    outputName = args.output

    # Load run data
    print('Loading data')
    resultPath = args.result
    runData = load(resultPath)
    
    catsVar = {
        'ref-noisy': [], 
        'thresholding': [],
        'search': [],
        'denoiser_coeff': []
    }

    # Generate rowData and sort by score
    print('Processing data')
    rowData = RowData(runData)
    if useBestMetrics:
        rowData.sort_row_data_metric()
        catsConst = lambda im : { 'imageLoader': [ im ] }
        results = [compute_stats(rowData, catsConst(im), catsVar, p, combsLength) for p in topPercentages for im in imageLoaders]
        outputName += '_bestMetrics'
    else:
        scoreIndex = RowData.HEADER.index('score')
        rowData.sort_row_data(scoreIndex)
        catsConst = lambda im, c: {
            'imageLoader': [ im ], 
            'metric': [ c ]
        }
        results = [compute_stats(rowData, catsConst(im, k), catsVar, p, combsLength) for p in topPercentages for im in imageLoaders for k in ['flip', 'hdrvdp3', 'mse', 'psnr', 'ssim']]

    print('Saving results')
    save(f'{outputName}.json', results, False)
    to_excel(f'{outputName}.xlsx', results)
    print('Done')

    # results = load('stats.json')
    # to_excel(results)

# The following code block will only execute if this script is run directly,
# not if it's imported as a module in another script.
if __name__ == "__main__":
    main()